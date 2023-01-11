#import openpyxl library to read excel file
from openpyxl import load_workbook
#load excel file
wb = load_workbook('test.xlsx')
#load sheet
ws = wb['Sheet1']
#ask input and find it in given column
search = input("Enter the value to search: ")
for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
    for cell in row:
        if cell.value == search:
            ws.delete_rows(cell.row) #delete row if found
            break
#save updated excel file
wb.save('test.xlsx')

