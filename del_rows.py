#import openpyxl library to read excel file
from openpyxl import load_workbook

#load excel file
wb1 = load_workbook('test.xlsx') #file to take reference
wb2 = load_workbook('test1.xlsx') #file to be updated

#load sheet
ws1 = wb1['Sheet1']
ws2 = wb2['Sheet1']

#ask input and find it in given column
for row1 in ws1.iter_rows(min_row=1, max_col=1, max_row=ws1.max_row):
    for cell1 in row1:
        for row2 in ws2.iter_rows(min_row=1, max_col=1, max_row=ws2.max_row):
            for cell2 in row2:
                 if cell1.value == cell2.value:
                     ws2.delete_rows(cell2.row) #delete row if found
                    
        
#save updated excel file
wb1.save('test.xlsx')
wb2.save('test1.xlsx')

